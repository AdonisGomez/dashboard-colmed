"""
Completa la columna de cuotas en Reporte_Montos-act.xlsx.
Para cada 'Monto total', si algún socio ya tiene el equivalente en cuotas definido,
se copia ese texto a todos los socios con el mismo monto que tengan la celda vacía.
"""
import pandas as pd
from openpyxl import load_workbook

ARCHIVO_ENTRADA = "Reporte_Montos-act.xlsx"
ARCHIVO_SALIDA = "Reporte_Montos-act_cuotas_completas.xlsx"
HOJA = "Montos por socio"

def main():
    # Leer datos (encabezado en fila 3 del Excel = índice 3)
    df = pd.read_excel(ARCHIVO_ENTRADA, sheet_name=HOJA, header=3)
    col_monto = "Monto total"
    col_cuotas = df.columns[3]  # columna de cuotas (Unnamed: 3 o similar)

    # Filas que ya tienen cuotas definidas (no vacías)
    mask_lleno = (
        df[col_cuotas].notna()
        & (df[col_cuotas].astype(str).str.strip() != "")
        & (df[col_cuotas].astype(str) != "nan")
    )
    con_cuotas = df.loc[mask_lleno]

    # Mapeo: cada monto -> texto de cuotas (primera aparición)
    monto_a_cuotas = con_cuotas.groupby(col_monto)[col_cuotas].first().to_dict()

    # Valores a escribir: para cada fila, si cuotas está vacía y hay mapeo, usar mapeo
    mask_vacio = (
        df[col_cuotas].isna()
        | (df[col_cuotas].astype(str).str.strip() == "")
        | (df[col_cuotas].astype(str) == "nan")
    )
    df["cuotas_nuevo"] = df[col_monto].map(monto_a_cuotas)
    a_rellenar = mask_vacio & df["cuotas_nuevo"].notna()

    # Abrir Excel con openpyxl para no tocar título ni formato
    wb = load_workbook(ARCHIVO_ENTRADA)
    ws = wb[HOJA]
    # Datos en Excel: fila 5 = primera fila de datos (fila 4 = encabezados)
    # pandas índice 0 -> Excel fila 5
    col_cuotas_excel = 4  # columna D

    count = 0
    for i in df.index:
        if a_rellenar.loc[i]:
            valor = df.loc[i, "cuotas_nuevo"]
            if pd.notna(valor) and str(valor).strip():
                excel_row = i + 5  # índice 0 -> fila 5
                ws.cell(row=excel_row, column=col_cuotas_excel, value=valor)
                count += 1

    wb.save(ARCHIVO_SALIDA)
    print(f"Listo. Se completaron {count} celdas de cuotas.")
    print(f"Montos con cuotas de referencia: {len(monto_a_cuotas)}.")
    print(f"Guardado en: {ARCHIVO_SALIDA}")

if __name__ == "__main__":
    main()
