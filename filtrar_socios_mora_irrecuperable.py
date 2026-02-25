"""
Filtra el archivo de montos por socio excluyendo socios con estado MI (mora irrecuperable)
según la base `BasesDeDatos-CUOTAS.xlsx`.

Entrada principal:
- sep/Reporte_Montos-act_cuotas_completas.xlsx  (hoja: Montos por socio)
- BasesDeDatos-CUOTAS.xlsx

Salidas:
- sep/Reporte_Montos_filtrado_sin_MI.xlsx        (mismos campos, sin socios MI)
- sep/socios_mora_irrecuperable.csv             (lista de socios excluidos con resumen)
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).parent
PATH_REPORTE = BASE_DIR / "sep" / "Reporte_Montos-act_cuotas_completas.xlsx"
PATH_CUOTAS = BASE_DIR / "BasesDeDatos-CUOTAS.xlsx"

OUT_REPORTE = BASE_DIR / "sep" / "Reporte_Montos_filtrado_sin_MI.xlsx"
OUT_SOCIOS_MI = BASE_DIR / "sep" / "socios_mora_irrecuperable.csv"


def cargar_reporte():
    df = pd.read_excel(PATH_REPORTE, sheet_name="Montos por socio", header=3)
    df.columns = [str(c).strip() for c in df.columns]

    # Detectar columnas por nombre aproximado
    col_socio = next(
        (c for c in df.columns if "código" in c.lower() or "codigo" in c.lower()),
        df.columns[0],
    )
    col_monto = next((c for c in df.columns if "monto total" in c.lower()), df.columns[1])
    col_fecha = next(
        (c for c in df.columns if "última fecha" in c.lower() or "ultima fecha" in c.lower()),
        df.columns[2],
    )
    col_cuotas = df.columns[3] if len(df.columns) > 3 else None

    return df, col_socio, col_monto, col_fecha, col_cuotas


def cargar_cuotas():
    df = pd.read_excel(PATH_CUOTAS)
    df.columns = [str(c).strip() for c in df.columns]
    # Columnas esperadas: socio_id, fecha_liquidacion, estado, anio, mes, monto, fecha_creacion, descripcion, ...
    col_socio = next((c for c in df.columns if c.lower() in ("socio_id", "codigo_socio", "codigo")), None)
    if col_socio is None:
        raise ValueError("No se encontró columna de socio en BasesDeDatos-CUOTAS.xlsx")
    col_estado = next((c for c in df.columns if "estado" in c.lower()), None)
    if col_estado is None:
        raise ValueError("No se encontró columna de estado en BasesDeDatos-CUOTAS.xlsx")
    col_monto = next((c for c in df.columns if c.lower() == "monto"), None)

    return df, col_socio, col_estado, col_monto


def main():
    # Cargar reporte de montos
    reporte, col_socio_rep, col_monto_rep, col_fecha_rep, col_cuotas_rep = cargar_reporte()

    # Cargar base de cuotas (con estados)
    cuotas, col_socio_cuotas, col_estado_cuotas, col_monto_cuotas = cargar_cuotas()

    # Normalizar tipos de socio a entero
    reporte[col_socio_rep] = pd.to_numeric(reporte[col_socio_rep], errors="coerce").astype("Int64")
    cuotas[col_socio_cuotas] = pd.to_numeric(cuotas[col_socio_cuotas], errors="coerce").astype("Int64")

    # Socios con estado MI (mora irrecuperable)
    mi_mask = cuotas[col_estado_cuotas].astype(str).str.upper().str.strip() == "MI"
    socios_mi = cuotas.loc[mi_mask, col_socio_cuotas].dropna().unique()
    socios_mi_set = set(int(x) for x in socios_mi)

    print(f"Socios con estado MI en BasesDeDatos-CUOTAS: {len(socios_mi_set)}")

    # Marcar cuáles socios del reporte están en MI
    reporte["es_MI"] = reporte[col_socio_rep].apply(lambda x: int(x) in socios_mi_set if pd.notna(x) else False)

    # DataFrame de socios MI presentes en el reporte (para revisar)
    socios_mi_en_reporte = (
        reporte.loc[reporte["es_MI"], [col_socio_rep, col_monto_rep]]
        .rename(columns={col_socio_rep: "socio_id", col_monto_rep: "monto_total_reporte"})
        .drop_duplicates()
    )

    # Resumen adicional: cuántos registros MI y monto total de MI en la base de cuotas
    if len(socios_mi_en_reporte) > 0:
        resumen_mi = (
            cuotas.loc[mi_mask & cuotas[col_socio_cuotas].isin(socios_mi_set)]
            .groupby(col_socio_cuotas)
            .agg(
                registros_MI=("estado", "count"),
                monto_total_MI=(col_monto_cuotas, "sum") if col_monto_cuotas else ("estado", "count"),
            )
            .reset_index()
            .rename(columns={col_socio_cuotas: "socio_id"})
        )
        socios_mi_en_reporte = socios_mi_en_reporte.merge(resumen_mi, on="socio_id", how="left")

    # Guardar lista de socios MI detectados
    OUT_SOCIOS_MI.parent.mkdir(parents=True, exist_ok=True)
    socios_mi_en_reporte.to_csv(OUT_SOCIOS_MI, index=False, encoding="utf-8-sig")
    print(f"Socios MI listados en: {OUT_SOCIOS_MI}")

    # Filtrar reporte quitando socios con mora irrecuperable
    filtrado = reporte[~reporte["es_MI"]].drop(columns=["es_MI"])
    print(f"Filas reporte original: {len(reporte)}")
    print(f"Filas después de quitar MI: {len(filtrado)}")

    # Crear nuevo Excel manteniendo las 4 primeras filas del original (título + encabezados)
    wb_orig = load_workbook(PATH_REPORTE)
    ws_orig = wb_orig["Montos por socio"]

    # Leer las primeras 4 filas (0: título, 1: total, 2: vacío, 3: encabezados)
    meta_rows = []
    for r in range(1, 5):
        meta_rows.append([cell.value for cell in ws_orig[r]])

    # Construir workbook nuevo
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Montos por socio"

    # Escribir filas meta
    for row in meta_rows:
        ws_new.append(row)

    # Asegurar orden de columnas como en el original
    cols = [col_socio_rep, col_monto_rep, col_fecha_rep]
    if col_cuotas_rep is not None:
        cols.append(col_cuotas_rep)

    for _, row in filtrado[cols].iterrows():
        ws_new.append(list(row.values))

    OUT_REPORTE.parent.mkdir(parents=True, exist_ok=True)
    wb_new.save(OUT_REPORTE)
    print(f"Reporte filtrado guardado en: {OUT_REPORTE}")


if __name__ == "__main__":
    main()

